# -*- coding: utf-8 -*-
"""
Tạo báo cáo Excel 1 FILE nhiều SHEET (tải 1 lần có hết).
Bố cục: Tổng quan (dashboard) | Bảng quét | Top 1M/3M/6M/YTD (mỗi sheet
tăng+giảm) | Sổ mua thử. Định dạng màu chuyên nghiệp.
Trả về bytes để st.download_button.
"""
from io import BytesIO
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

FONT = "Arial"
NAVY = "1F2A44"; BLUE = "2962FF"; GREEN = "1B7F4B"; RED = "C0392B"
LIGHT = "EAF1FB"; GREY = "F2F2F2"; WHITE = "FFFFFF"
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _hdr(ws, row, col, text, fill=NAVY, color=WHITE, size=11, bold=True):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name=FONT, bold=bold, color=color, size=size)
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER
    return c


def _title(ws, row, text, span=6, size=14):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, bold=True, color=NAVY, size=size)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    return c


def _write_table(ws, start_row, df, pnl_col=None):
    """Ghi DataFrame vào ws bắt đầu start_row. Trả row kế tiếp."""
    if df is None or df.empty:
        ws.cell(row=start_row, column=1, value="(không có dữ liệu)").font = Font(name=FONT, italic=True)
        return start_row + 2
    cols = list(df.columns)
    for j, name in enumerate(cols, 1):
        _hdr(ws, start_row, j, str(name))
    r = start_row + 1
    for _, row in df.iterrows():
        for j, name in enumerate(cols, 1):
            v = row[name]
            cell = ws.cell(row=r, column=j, value=(None if pd.isna(v) else v))
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if j > 1 else "left")
            # tô màu cột % lời/lỗ hoặc % thay đổi
            if pnl_col and name == pnl_col and pd.notna(v):
                try:
                    fv = float(v)
                    cell.font = Font(name=FONT, size=10, bold=True,
                                     color=GREEN if fv >= 0 else RED)
                except Exception:
                    pass
        r += 1
    return r + 1


def _autosize(ws, max_w=40):
    for col in ws.columns:
        length = 8
        letter = None
        for cell in col:
            if letter is None and hasattr(cell, "column_letter"):
                letter = cell.column_letter
            try:
                if cell.value is not None:
                    length = max(length, len(str(cell.value)) + 2)
            except Exception:
                pass
        if letter:
            ws.column_dimensions[letter].width = min(length, max_w)


def _movers_tables(df, period):
    """Trả (df_tang, df_giam) cho 1 mốc."""
    if period not in df.columns:
        return None, None
    v = df[df[period].notna()].copy()
    v[period] = pd.to_numeric(v[period], errors="coerce")
    v = v.dropna(subset=[period])
    show = [c for c in ["Mã", "Ngành", "Trạng thái", period, "Vùng chờ", "RSI", "Giá"] if c in v.columns]
    up = v.sort_values(period, ascending=False).head(15)[show]
    down = v.sort_values(period, ascending=True).head(15)[show]
    return up, down


def build_report(rows, meta=None, portfolio=None, conclusion=None, seasonality=None):
    """rows: list dict từ quét. portfolio: list dict (đã có pnl) hoặc None.
    conclusion: dict từ insights.market_conclusion. seasonality: dict từ insights.seasonality.
    Trả bytes .xlsx."""
    meta = meta or {}
    df = pd.DataFrame(rows) if rows else pd.DataFrame()
    wb = Workbook()

    # ---------- Sheet 1: Tổng quan ----------
    ws = wb.active; ws.title = "Tổng quan"
    ws.sheet_view.showGridLines = False
    _title(ws, 1, "📊 BÁO CÁO QUÉT THỊ TRƯỜNG — VN STOCK ANALYST DESK", span=6, size=15)
    ws.cell(row=2, column=1, value=f"Thời điểm: {datetime.now():%Y-%m-%d %H:%M}").font = Font(name=FONT, italic=True, color="666666")

    r = 4
    if not df.empty and "status" in df.columns:
        n_vao = int((df["status"] == "VAO").sum())
        n_cho = int((df["status"] == "CHO").sum())
        n_tranh = int((df["status"] == "TRANH").sum()) if "TRANH" in df["status"].values else 0
        n_theodoi = len(df) - n_vao - n_cho - n_tranh
        cards = [("Tổng số mã", len(df), BLUE), ("🟢 Điểm vào", n_vao, GREEN),
                 ("🟡 Chờ", n_cho, "B8860B"), ("🔴 Tránh", n_tranh, RED),
                 ("🔵 Theo dõi", n_theodoi, NAVY)]
        for i, (label, val, color) in enumerate(cards):
            col = 1 + i
            lc = ws.cell(row=r, column=col, value=label)
            lc.font = Font(name=FONT, bold=True, color=WHITE, size=10)
            lc.fill = PatternFill("solid", fgColor=color)
            lc.alignment = Alignment(horizontal="center")
            lc.border = BORDER
            vc = ws.cell(row=r+1, column=col, value=val)
            vc.font = Font(name=FONT, bold=True, size=16, color=color)
            vc.alignment = Alignment(horizontal="center")
            vc.border = BORDER
        r += 3

    # thống kê tăng/giảm theo mốc
    _title(ws, r, "Thống kê biến động (số mã tăng / giảm)", span=6, size=12); r += 1
    period_labels = [("%1M", "1 tháng"), ("%3M", "3 tháng"), ("%6M", "6 tháng"), ("%YTD", "Từ đầu năm")]
    _hdr(ws, r, 1, "Mốc"); _hdr(ws, r, 2, "Số mã TĂNG"); _hdr(ws, r, 3, "Số mã GIẢM")
    _hdr(ws, r, 4, "Tăng mạnh nhất"); _hdr(ws, r, 5, "Giảm sâu nhất")
    r += 1
    for pcol, plabel in period_labels:
        if pcol in df.columns:
            v = pd.to_numeric(df[pcol], errors="coerce").dropna()
            n_up = int((v > 0).sum()); n_down = int((v < 0).sum())
            top_up = f"{v.max():+.1f}%" if len(v) else "—"
            top_down = f"{v.min():+.1f}%" if len(v) else "—"
            ws.cell(row=r, column=1, value=plabel).font = Font(name=FONT, bold=True)
            ws.cell(row=r, column=2, value=n_up).font = Font(name=FONT, color=GREEN, bold=True)
            ws.cell(row=r, column=3, value=n_down).font = Font(name=FONT, color=RED, bold=True)
            ws.cell(row=r, column=4, value=top_up)
            ws.cell(row=r, column=5, value=top_down)
            for cc in range(1, 6):
                ws.cell(row=r, column=cc).border = BORDER
                ws.cell(row=r, column=cc).alignment = Alignment(horizontal="center")
            r += 1
    if meta.get("liq"):
        r += 1
        ws.cell(row=r, column=1, value=f"Thanh khoản HOSE (ước): {meta['liq']:,.0f} tỷ").font = Font(name=FONT, italic=True)
    r += 2

    # Kết luận tổng hợp
    if conclusion:
        _title(ws, r, f"🧭 KẾT LUẬN: {conclusion.get('verdict','')}", span=6, size=12); r += 1
        for ln in conclusion.get("lines", []):
            ws.cell(row=r, column=1, value="• " + ln).font = Font(name=FONT, size=10)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
        act = ws.cell(row=r, column=1, value="👉 " + conclusion.get("action", ""))
        act.font = Font(name=FONT, size=10, bold=True, color=BLUE)
        act.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=6)
        ws.row_dimensions[r].height = 50; r += 4
    r += 1
    ws.cell(row=r, column=1, value="Tham khảo, KHÔNG phải khuyến nghị đầu tư. Đầu tư tự chịu trách nhiệm.").font = Font(name=FONT, italic=True, color="999999", size=9)
    _autosize(ws)

    # ---------- Sheet 2: Bảng quét ----------
    ws2 = wb.create_sheet("Bảng quét")
    ws2.sheet_view.showGridLines = False
    _title(ws2, 1, "Bảng quét đầy đủ", span=8)
    if not df.empty:
        cols = [c for c in ["Mã","Trạng thái","Vùng","Vùng chờ","Ngành","Gió ngành","Điểm CB",
                "Xu hướng","RSI","%1M","%3M","%6M","%YTD","P/E","ROE%","Cảnh báo","Giá","Setup"]
                if c in df.columns]
        _write_table(ws2, 3, df[cols])
    _autosize(ws2, max_w=50)

    # ---------- Sheet 3-6: Top theo mốc ----------
    for pcol, plabel in period_labels:
        wsm = wb.create_sheet(f"Top {plabel}")
        wsm.sheet_view.showGridLines = False
        up, down = _movers_tables(df, pcol)
        _title(wsm, 1, f"🔺 Top TĂNG mạnh — {plabel}", span=7, size=12)
        nxt = _write_table(wsm, 2, up, pnl_col=pcol)
        _title(wsm, nxt + 1, f"🔻 Top GIẢM sâu — {plabel} (soi điểm vào)", span=7, size=12)
        _write_table(wsm, nxt + 2, down, pnl_col=pcol)
        _autosize(wsm)

    # ---------- Sheet: Mùa vụ ----------
    if seasonality and seasonality.get("monthly"):
        wss = wb.create_sheet("Mùa vụ")
        wss.sheet_view.showGridLines = False
        _title(wss, 1, "📅 Mùa vụ VN-Index — lợi suất TB theo tháng", span=4, size=13)
        wss.cell(row=2, column=1, value=seasonality.get("note", "")).font = Font(name=FONT, italic=True, color="999999", size=9)
        _hdr(wss, 4, 1, "Tháng"); _hdr(wss, 4, 2, "Lợi suất TB %")
        _hdr(wss, 4, 3, "Tỷ lệ tăng %"); _hdr(wss, 4, 4, "Số năm")
        rr = 5
        VN_MO = ["", "Tháng 1","Tháng 2","Tháng 3","Tháng 4","Tháng 5","Tháng 6",
                 "Tháng 7","Tháng 8","Tháng 9","Tháng 10","Tháng 11","Tháng 12"]
        for mo in range(1, 13):
            mm = seasonality["monthly"].get(mo)
            if mm:
                wss.cell(row=rr, column=1, value=VN_MO[mo]).font = Font(name=FONT, bold=True)
                vc = wss.cell(row=rr, column=2, value=mm["avg"])
                vc.font = Font(name=FONT, bold=True, color=GREEN if mm["avg"] >= 0 else RED)
                wss.cell(row=rr, column=3, value=mm["win_rate"])
                wss.cell(row=rr, column=4, value=mm["n"])
                for cc in range(1, 5):
                    wss.cell(row=rr, column=cc).border = BORDER
                    wss.cell(row=rr, column=cc).alignment = Alignment(horizontal="center" if cc > 1 else "left")
                rr += 1
        _autosize(wss)

    # ---------- Sheet: Sổ mua thử ----------
    if portfolio:
        wsp = wb.create_sheet("Sổ mua thử")
        wsp.sheet_view.showGridLines = False
        _title(wsp, 1, "💼 Sổ mua thử (paper trading) — lời/lỗ", span=8, size=13)
        dfp = pd.DataFrame(portfolio).rename(columns={
            "sym":"Mã","entry":"Giá mua","current":"Giá hiện tại","qty":"SL",
            "pnl_pct":"Lời/Lỗ %","pnl_value":"Lời/Lỗ (ngđ)","date":"Ngày mua","note":"Ghi chú"})
        pcols = [c for c in ["Mã","Ngày mua","Giá mua","Giá hiện tại","SL","Lời/Lỗ %","Lời/Lỗ (ngđ)","Ghi chú"] if c in dfp.columns]
        _write_table(wsp, 3, dfp[pcols], pnl_col="Lời/Lỗ %")
        _autosize(wsp)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()
